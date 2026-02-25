from __future__ import annotations

import asyncio
import hashlib
import json
import re
from datetime import datetime
from html import escape
from pathlib import Path
from typing import Any, Dict, List, Tuple

import aiohttp
import gradio as gr
import pandas as pd
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.page import PageMargins

from config_manager import LLMConfigManager
from extractor_core import VocabularySentenceExtractor
from llm_processor import (
    AsyncRunner,
    CheckpointManager,
    FatalAPIError,
    LLMClient,
    TransientAPIError,
)


SETTINGS_PATH = Path("llm_settings.json")
CHECKPOINT_PATH = Path("cache_results.jsonl")
TRACE_LOG_PATH = Path("run_trace.log")
BUILTIN_DICT_DIR = Path("dicts")
BUILTIN_IELTS_CSV = BUILTIN_DICT_DIR / "ielts_words.csv"
BUILTIN_BLACKLIST_CSV = BUILTIN_DICT_DIR / "common_blacklist.csv"
CONFIG_MANAGER = LLMConfigManager(SETTINGS_PATH)


def ensure_builtin_dict_assets() -> None:
    """初始化内置词库资产，避免容器首次启动缺文件。"""
    BUILTIN_DICT_DIR.mkdir(parents=True, exist_ok=True)
    if not BUILTIN_IELTS_CSV.exists():
        BUILTIN_IELTS_CSV.write_text(
            "word,level\n"
            "accommodate,5\n"
            "aberration,7\n"
            "mumble,6\n"
            "meticulous,7\n"
            "apple,1\n",
            encoding="utf-8",
        )
    if not BUILTIN_BLACKLIST_CSV.exists():
        BUILTIN_BLACKLIST_CSV.write_text(
            "word\nthe\nof\nand\nto\na\n",
            encoding="utf-8",
        )


ensure_builtin_dict_assets()


API_TYPE_TO_LABEL = {
    "openai_official": "OpenAI官方",
    "openai_compatible": "兼容OpenAI/第三方",
    "openai_responses": "兼容OpenAI (特殊Responses协议)",
    "gemini_official": "Gemini官方",
    "ollama_local": "Ollama本地",
}
LABEL_TO_API_TYPE = {v: k for k, v in API_TYPE_TO_LABEL.items()}


def now_text() -> str:
    return datetime.now().strftime("%H:%M:%S")


def calculate_file_hash(file_path: str, algorithm: str = "sha1") -> str:
    """流式计算文件哈希，避免一次性读入大文件。"""
    if algorithm.lower() != "sha1":
        raise ValueError("当前仅支持 sha1")

    hasher = hashlib.sha1()
    with open(file_path, "rb") as f:
        while True:
            chunk = f.read(8192)
            if not chunk:
                break
            hasher.update(chunk)
    return hasher.hexdigest()


def parse_provider_model(text: str) -> Tuple[str, str]:
    value = (text or "").strip()
    if ":" not in value:
        return "", ""
    provider, model = value.split(":", 1)
    return provider.strip(), model.strip()


def get_task_choices() -> Tuple[List[str], str]:
    pairs = CONFIG_MANAGER.get_provider_model_pairs()
    settings = CONFIG_MANAGER.load_settings()
    choices = [f"{provider}:{model}" for provider, model in pairs]

    default_provider = str(settings.get("default_provider", "")).strip()
    default_model = str(settings.get("default_model", "")).strip()
    default_choice = (
        f"{default_provider}:{default_model}"
        if default_provider and default_model
        else ""
    )

    if default_choice not in choices:
        default_choice = choices[0] if choices else ""
    return choices, default_choice


def get_provider_selector_choices() -> List[str]:
    return ["新增提供商", *CONFIG_MANAGER.get_provider_names()]


def map_extract_mode(label: str) -> str:
    text = (label or "").strip()
    if text == "仅识别单词":
        return "word_only"
    if text == "仅识别固定搭配":
        return "collocation_only"
    return "both"


def resolve_page_orientation(page_orientation: str, include_sentence: bool) -> str:
    choice = (page_orientation or "").strip()
    if choice == "竖向":
        return "portrait"
    if choice == "横向":
        return "landscape"
    return "landscape" if include_sentence else "portrait"


def summarize_output_formats(selected_formats: List[str]) -> str:
    selected = selected_formats or []
    if not selected:
        return "**本次将导出**：未选择格式（将回退为 Excel）"
    return f"**本次将导出**：{', '.join(selected)}"


def load_results_dataframe(
    cache_path: Path, word_sentence_map: Dict[str, str]
) -> pd.DataFrame:
    if not cache_path.exists():
        return pd.DataFrame(
            columns=[
                "word",
                "part_of_speech",
                "context_meaning",
                "collocation",
                "collocation_meaning",
                "sentence",
            ]
        )

    rows: List[Dict[str, Any]] = []
    with cache_path.open("r", encoding="utf-8") as f:
        for raw_line in f:
            line = raw_line.strip()
            if not line:
                continue
            try:
                item = json.loads(line)
            except json.JSONDecodeError:
                continue

            word = str(item.get("word", "")).strip()
            part_of_speech = str(item.get("part_of_speech", "")).strip()
            context_meaning = str(item.get("context_meaning", "")).strip()
            collocation = str(item.get("collocation", "")).strip()
            collocation_meaning = str(item.get("collocation_meaning", "")).strip()
            sentence = str(item.get("sentence", "")).strip() or word_sentence_map.get(
                word.lower(), ""
            )

            # 与解析层保持一致：搭配占位词统一视为空值，保证“单词优先”。
            empty_markers = {
                "none",
                "null",
                "n/a",
                "no collocation",
                "na",
                "nil",
                "无",
                "没有",
            }
            if collocation.lower() in empty_markers:
                collocation = ""
            if collocation_meaning.lower() in empty_markers:
                collocation_meaning = ""
            if not collocation:
                collocation_meaning = ""

            if word:
                rows.append(
                    {
                        "word": word,
                        "part_of_speech": part_of_speech,
                        "context_meaning": context_meaning,
                        "collocation": collocation,
                        "collocation_meaning": collocation_meaning,
                        "sentence": sentence,
                    }
                )

    return pd.DataFrame(rows)


def _generate_pdf(
    excel_df: pd.DataFrame,
    pdf_path: Path,
    include_sentence: bool,
    color_theme: str,
    page_orientation: str,
) -> None:
    """使用 ReportLab 生成 PDF（A4，方向可配置，中文可读，强清洗）。"""
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.cidfonts import UnicodeCIDFont
    from reportlab.platypus import Paragraph, SimpleDocTemplate, Table, TableStyle

    def _sanitize_pdf_text(value: Any) -> str:
        if value is None:
            return ""
        try:
            if pd.isna(value):
                return ""
        except Exception:
            pass
        text = str(value)
        # 清理控制字符并转义 ReportLab 段落标记，避免生成异常或损坏内容流。
        text = re.sub(r"[\x00-\x08\x0B\x0C\x0E-\x1F]", "", text)
        text = text.strip().strip('"').strip("'")
        text = text.replace("\r\n", "\n").replace("\r", "\n").replace("\\n", "\n")
        text = escape(text, quote=False)
        return text.replace("\n", "<br/>")

    pdfmetrics.registerFont(UnicodeCIDFont("STSong-Light"))

    body_style = ParagraphStyle(
        "Body",
        fontName="STSong-Light",
        fontSize=10,
        leading=13,
        wordWrap="CJK",
    )
    header_style = ParagraphStyle(
        "Header",
        fontName="STSong-Light",
        fontSize=12,
        leading=14,
        alignment=1,  # center
        wordWrap="CJK",
    )

    headers = [str(c) for c in excel_df.columns]

    # 固定绝对列宽（points）
    base_widths = {
        "单词": 80,
        "词性": 40,
        "语境释义": 100,
        "固定搭配": 120,
        "搭配释义": 120,
        "原句": 340,
    }

    if include_sentence and "原句" in headers:
        col_widths = [base_widths.get(h, 100) for h in headers]
    else:
        # 无原句：将 340 的剩余宽度平均分给“释义/搭配”列
        expandable = [h for h in headers if h in {"语境释义", "固定搭配", "搭配释义"}]
        bonus = 340 // len(expandable) if expandable else 0
        col_widths = []
        for h in headers:
            width = base_widths.get(h, 100)
            if h in expandable:
                width += bonus
            if h == "原句":
                width = 0
            col_widths.append(width)

    pdf_orientation = resolve_page_orientation(page_orientation, include_sentence)
    pagesize = landscape(A4) if pdf_orientation == "landscape" else A4
    left_margin = 18
    right_margin = 18
    available_width = pagesize[0] - left_margin - right_margin
    total_width = sum(col_widths)
    if total_width > 0 and total_width > available_width:
        scale = available_width / total_width
        col_widths = [round(w * scale, 2) for w in col_widths]

    # 构建表格数据
    data: List[List[Paragraph]] = []
    data.append(
        [Paragraph(f"<b>{_sanitize_pdf_text(h)}</b>", header_style) for h in headers]
    )
    for row in excel_df.itertuples(index=False):
        cleaned_row: List[Paragraph] = []
        for cell in row:
            cleaned_row.append(Paragraph(_sanitize_pdf_text(cell), body_style))
        data.append(cleaned_row)

    tmp_pdf_path = pdf_path.with_suffix(f"{pdf_path.suffix}.tmp")
    if tmp_pdf_path.exists():
        tmp_pdf_path.unlink()

    doc = SimpleDocTemplate(
        str(tmp_pdf_path),
        pagesize=pagesize,
        leftMargin=left_margin,
        rightMargin=right_margin,
        topMargin=18,
        bottomMargin=18,
    )

    table = Table(data, colWidths=col_widths, repeatRows=1)

    style_cmds = [
        ("FONTNAME", (0, 0), (-1, -1), "STSong-Light"),
        ("FONTSIZE", (0, 0), (-1, 0), 12),
        ("FONTSIZE", (0, 1), (-1, -1), 10),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.black),
        ("LEFTPADDING", (0, 0), (-1, -1), 4),
        ("RIGHTPADDING", (0, 0), (-1, -1), 4),
    ]

    # 对齐：原句左对齐，其余居中
    style_cmds.append(("ALIGN", (0, 0), (-1, -1), "CENTER"))
    if "原句" in headers:
        sentence_idx = headers.index("原句")
        style_cmds.append(("ALIGN", (sentence_idx, 0), (sentence_idx, -1), "LEFT"))

    if color_theme == "护眼晨雾 (低饱和彩)":
        style_cmds.append(("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#E7F3F6")))
        for row_idx in range(1, len(data)):
            if row_idx % 2 == 0:
                style_cmds.append(
                    (
                        "BACKGROUND",
                        (0, row_idx),
                        (-1, row_idx),
                        colors.HexColor("#F9F9F9"),
                    )
                )

    table.setStyle(TableStyle(style_cmds))
    try:
        doc.build([table])
        tmp_pdf_path.replace(pdf_path)
    finally:
        if tmp_pdf_path.exists():
            tmp_pdf_path.unlink()


def export_files(
    cache_path: Path,
    word_sentence_map: Dict[str, str],
    extract_mode: str,
    include_meaning: bool,
    include_sentence: bool,
    color_theme: str,
    page_orientation: str,
    output_formats: List[str],
) -> List[str]:
    def _sanitize_excel_text(value: Any) -> Any:
        if value is None:
            return ""
        text = str(value)
        # 去掉 Excel/OpenXML 非法控制字符，避免 IllegalCharacterError
        return re.sub(r"[\x00-\x08\x0B\x0C\x0E-\x1F]", "", text)

    df = load_results_dataframe(cache_path, word_sentence_map)

    excel_cols = ["word", "part_of_speech"]
    if include_meaning:
        excel_cols.append("context_meaning")
    if extract_mode != "word_only":
        excel_cols.extend(["collocation", "collocation_meaning"])
    if include_sentence:
        excel_cols.append("sentence")

    for col in excel_cols:
        if col not in df.columns:
            df[col] = ""

    excel_df = df[excel_cols].rename(
        columns={
            "word": "单词",
            "part_of_speech": "词性",
            "context_meaning": "语境释义",
            "collocation": "固定搭配",
            "collocation_meaning": "搭配释义",
            "sentence": "原句",
        }
    )
    excel_df = excel_df.map(_sanitize_excel_text)

    reading_excel = Path("vocabulary_reading.xlsx")
    typing_csv = Path("typing_world.csv")

    with pd.ExcelWriter(reading_excel, engine="openpyxl") as writer:
        excel_df.to_excel(writer, sheet_name="Sheet1", index=False)
        ws = writer.sheets["Sheet1"]

        # A4 打印适配：支持自动/竖向/横向切换
        ws.page_setup.paperSize = 9  # A4
        ws.page_setup.orientation = resolve_page_orientation(
            page_orientation, include_sentence
        )
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0
        ws.page_margins = PageMargins(left=0.75, right=0.75, top=0.75, bottom=0.75)
        ws.print_title_rows = "1:1"

        # 列宽：按列语义映射，避免动态列错位
        width_map = {
            "单词": 18,
            "词性": 10,
            "语境释义": 25,
            "固定搭配": 25,
            "搭配释义": 25,
            "原句": 60,
        }
        header_by_col: Dict[int, str] = {}
        for col_idx in range(1, ws.max_column + 1):
            header = str(ws.cell(row=1, column=col_idx).value or "").strip()
            header_by_col[col_idx] = header
            col_letter = ws.cell(row=1, column=col_idx).column_letter
            ws.column_dimensions[col_letter].width = width_map.get(header, 20)

        # 字体/对齐/边框（黑白高对比）
        header_font = Font(name="Arial", size=14, bold=True, color="000000")
        body_font = Font(name="Arial", size=12, bold=False, color="000000")
        align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
        align_sentence = Alignment(horizontal="left", vertical="center", wrap_text=True)
        border_color = "000000" if color_theme == "黑白公文 (打印专用)" else "BFBFBF"
        thin_side = Side(style="thin", color=border_color)
        thin_border = Border(
            left=thin_side, right=thin_side, top=thin_side, bottom=thin_side
        )

        header_fill = PatternFill(fill_type=None)
        even_row_fill = PatternFill(fill_type=None)
        if color_theme == "护眼晨雾 (低饱和彩)":
            header_fill = PatternFill(fill_type="solid", fgColor="E7F3F6")
            even_row_fill = PatternFill(fill_type="solid", fgColor="F9F9F9")

        for row_idx in range(1, ws.max_row + 1):
            for col_idx in range(1, ws.max_column + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                if header_by_col.get(col_idx) == "原句":
                    cell.alignment = align_sentence
                else:
                    cell.alignment = align_center
                cell.border = thin_border
                if row_idx == 1:
                    cell.font = header_font
                    cell.fill = header_fill
                else:
                    cell.font = body_font
                    if color_theme == "护眼晨雾 (低饱和彩)" and row_idx % 2 == 0:
                        cell.fill = even_row_fill

    typing_source_col = "context_meaning"
    if typing_source_col not in df.columns:
        df[typing_source_col] = ""
    typing_df = df[["word", typing_source_col]].fillna("")
    typing_df = typing_df.map(_sanitize_excel_text)
    typing_df.to_csv(typing_csv, index=False, header=False, sep=",")

    output_files: List[str] = []
    if "Excel (.xlsx)" in output_formats:
        output_files.append(str(reading_excel.resolve()))

    if "PDF (.pdf)" in output_formats:
        pdf_path = Path("vocabulary_reading.pdf")
        try:
            _generate_pdf(
                excel_df,
                pdf_path,
                include_sentence=include_sentence,
                color_theme=color_theme,
                page_orientation=page_orientation,
            )
            output_files.append(str(pdf_path.resolve()))
        except Exception as exc:
            print(f"[Warning] PDF 生成失败，已跳过，仅保留 Excel: {exc}")
            if pdf_path.exists():
                try:
                    pdf_path.unlink()
                except OSError:
                    pass
            if str(reading_excel.resolve()) not in output_files:
                output_files.append(str(reading_excel.resolve()))

    if "TXT 单词本 (适配墨墨/欧路)" in output_formats:
        txt_path = Path("maimemo_vocabulary.txt")
        words = (
            df["word"]
            .fillna("")
            .astype(str)
            .str.strip()
            .str.lower()
            .drop_duplicates()
            .tolist()
        )
        words = [w for w in words if w]
        txt_path.write_text("\n".join(words), encoding="utf-8")
        output_files.append(str(txt_path.resolve()))

    if not output_files:
        output_files.append(str(reading_excel.resolve()))

    return output_files


async def process_pipeline(
    provider_model: str,
    min_level: float,
    extract_mode_label: str,
    include_meaning: bool,
    include_sentence: bool,
    color_theme: str,
    page_orientation: str,
    output_formats: List[str],
    filter_strategy_label: str,
    pdf_file: str,
    custom_csv_file: str,
):
    logs: List[str] = []

    def push(msg: str) -> str:
        line = f"[{now_text()}] {msg}"
        logs.append(line)
        with TRACE_LOG_PATH.open("a", encoding="utf-8") as f:
            f.write(line + "\n")
        return "\n".join(logs)

    try:
        if not pdf_file:
            yield push("请先上传电子书文件。"), []
            return

        provider_name, model_name = parse_provider_model(provider_model)
        if not provider_name or not model_name:
            yield push("请选择有效的“提供商:模型”配置。"), []
            return

        filename = Path(pdf_file).name
        file_hash = calculate_file_hash(pdf_file, algorithm="sha1")
        book_key = f"{file_hash[:8]}_{filename}"
        yield push(f"文件指纹: {book_key}"), []

        provider = CONFIG_MANAGER.get_provider(provider_name)
        if not provider:
            yield (
                push(f"配置不存在：{provider_name}。请在“模型配置中心”保存后重试。"),
                [],
            )
            return

        api_type = (
            str(provider.get("api_type", "openai_compatible")).strip()
            or "openai_compatible"
        )
        base_url = str(provider.get("base_url", "")).strip()
        api_key = str(provider.get("api_key", "")).strip()
        if not base_url:
            yield push(f"提供商 {provider_name} 缺少 Base URL，请到配置中心补全。"), []
            return

        extract_mode = map_extract_mode(extract_mode_label)
        yield (
            push(
                f"开始任务，使用配置={provider_name}:{model_name}，最低等级={int(min_level)}，提取模式={extract_mode_label}，页面方向={page_orientation}"
            ),
            [],
        )

        csv_path: str | None = None
        filter_strategy = "whitelist"
        effective_min_level = float(min_level)
        if custom_csv_file:
            csv_path = custom_csv_file
            yield push("检测到自定义词库：将覆盖内置词库配置。"), []

            try:
                custom_df = pd.read_csv(csv_path, encoding="utf-8-sig")
                has_level = "level" in custom_df.columns
            except Exception:
                has_level = False

            file_name_lower = Path(csv_path).name.lower()
            if has_level:
                filter_strategy = "whitelist"
                yield push("检测到自定义词库包含 level 列，自动使用白名单策略。"), []
            elif "black" in file_name_lower:
                filter_strategy = "blacklist"
                effective_min_level = 0
                yield push("检测到自定义黑名单词库文件名，自动使用黑名单策略。"), []
            else:
                filter_strategy = "whitelist"
                yield push("未识别到黑名单特征，默认按白名单词库处理。"), []
        else:
            if filter_strategy_label == "雅思备考模式 (白名单)":
                if BUILTIN_IELTS_CSV.exists():
                    csv_path = str(BUILTIN_IELTS_CSV)
                    yield push(f"使用内置词库：{BUILTIN_IELTS_CSV}"), []
                else:
                    yield push(f"任务失败: 未找到内置词库文件 {BUILTIN_IELTS_CSV}"), []
                    return
            else:
                if BUILTIN_BLACKLIST_CSV.exists():
                    csv_path = str(BUILTIN_BLACKLIST_CSV)
                    filter_strategy = "blacklist"
                    effective_min_level = 0
                    yield push(f"使用内置黑名单：{BUILTIN_BLACKLIST_CSV}"), []
                else:
                    yield (
                        push(f"任务失败: 未找到内置黑名单文件 {BUILTIN_BLACKLIST_CSV}"),
                        [],
                    )
                    return

        extractor = VocabularySentenceExtractor()
        if filter_strategy == "whitelist":
            yield push("正在按白名单词库与等级条件进行过滤..."), []
            extractor.load_vocabulary_from_csv(csv_path, min_level=effective_min_level)
            progress_stream = extractor.extract_with_progress(
                pdf_file,
                filter_strategy="whitelist",
                max_extracted_words=0,
                progress_every_sentences=100,
            )
        else:
            yield push("正在按黑名单策略提取专业词...（已忽略最低等级）"), []
            extractor.load_blacklist_from_csv(csv_path)
            progress_stream = extractor.extract_with_progress(
                pdf_file,
                filter_strategy="blacklist",
                max_extracted_words=0,
                progress_every_sentences=100,
            )

        word_sentence_map: Dict[str, str] = {}
        parse_done_event: Dict[str, Any] = {}
        for event in progress_stream:
            status = str(event.get("status", "")).strip()
            if status == "parsing":
                message = str(event.get("message", "")).strip()
                if message:
                    yield push(message), []
                continue
            if status == "done":
                parse_done_event = event
                data = event.get("data", {})
                if isinstance(data, dict):
                    word_sentence_map = data
                break

        if parse_done_event.get("truncated"):
            limit = parse_done_event.get("limit", 0)
            yield push(f"解析已触发上限保护：最多提取 {limit} 个候选词。"), []

        yield push(f"电子书解析完成，命中去重词汇数={len(word_sentence_map)}"), []

        checkpoint = CheckpointManager(CHECKPOINT_PATH)
        processed = checkpoint.load_processed_words()
        pending_items = checkpoint.filter_pending(
            word_sentence_map, book_key=book_key, extract_mode=extract_mode
        )
        yield (
            push(f"断点恢复完成，已处理={len(processed)}，待处理={len(pending_items)}"),
            [],
        )

        client = LLMClient(
            api_type=api_type,
            base_url=base_url,
            api_key=api_key,
            model_name=model_name,
        )

        yield push("正在进行预检（Preflight）..."), []
        try:
            async with aiohttp.ClientSession(
                timeout=aiohttp.ClientTimeout(total=30)
            ) as preflight_session:
                await client.preflight_check(preflight_session)
        except FatalAPIError as exc:
            yield (
                push(
                    f"【严重错误：预检失败，任务终止！请检查模型配置、API Key 或接口协议是否正确】{exc}"
                ),
                [],
            )
            return
        except TransientAPIError as exc:
            yield push(f"【预检失败：网络或服务暂时不可用】{exc}"), []
            return

        runner = AsyncRunner(client, checkpoint, max_concurrency=10)
        progress_queue: asyncio.Queue[str] = asyncio.Queue()

        yield push("开始请求 LLM，并发池启动（Semaphore=10）..."), []
        run_task = asyncio.create_task(
            runner.run(
                pending_items,
                progress_queue=progress_queue,
                book_key=book_key,
                extract_mode=extract_mode,
            )
        )

        while True:
            if run_task.done() and progress_queue.empty():
                break
            try:
                progress_line = await asyncio.wait_for(
                    progress_queue.get(), timeout=0.3
                )
                yield push(progress_line), []
            except asyncio.TimeoutError:
                if run_task.done():
                    break

        results = await run_task
        yield push(f"LLM 阶段完成，成功处理={len(results)}"), []

        yield push("正在导出 Excel 和 Typing-World CSV..."), []
        output_paths = export_files(
            CHECKPOINT_PATH,
            word_sentence_map,
            extract_mode,
            include_meaning=include_meaning,
            include_sentence=include_sentence,
            color_theme=color_theme,
            page_orientation=page_orientation,
            output_formats=output_formats,
        )
        yield push("导出完成，可下载结果文件。"), output_paths

    except asyncio.CancelledError:
        yield push("任务已停止。"), []
        raise
    except FatalAPIError as exc:
        yield push(f"【严重错误】{exc}，任务已终止。"), []
    except TransientAPIError as exc:
        yield push(f"【临时错误】{exc}，请稍后重试。"), []
    except RuntimeError as exc:
        if "连续失败超过阈值" in str(exc):
            yield push(f"【熔断触发】{exc}"), []
            return
        yield push(f"任务失败: {exc}"), []
    except Exception as exc:
        yield push(f"任务失败: {exc}"), []


def on_provider_for_edit_change(provider_for_edit: str):
    if provider_for_edit == "新增提供商":
        return "", "兼容OpenAI/第三方", "", "", gr.update(choices=[], value=None), ""

    provider = CONFIG_MANAGER.get_provider(provider_for_edit)
    if not provider:
        return (
            "",
            "兼容OpenAI/第三方",
            "",
            "",
            gr.update(choices=[], value=None),
            "未找到该提供商配置",
        )

    api_type = str(provider.get("api_type", "openai_compatible"))
    api_label = API_TYPE_TO_LABEL.get(api_type, "兼容OpenAI/第三方")
    base_url = str(provider.get("base_url", ""))
    api_key = str(provider.get("api_key", ""))
    models = [str(m).strip() for m in provider.get("models", []) if str(m).strip()]
    model_value = models[0] if models else None

    return (
        provider_for_edit,
        api_label,
        base_url,
        api_key,
        gr.update(choices=models, value=model_value),
        "",
    )


async def on_fetch_models(api_type_label: str, base_url: str, api_key: str):
    api_type = LABEL_TO_API_TYPE.get(api_type_label, "openai_compatible")
    models, err = await CONFIG_MANAGER.fetch_models(api_type, base_url, api_key)
    if err:
        gr.Warning(f"拉取模型失败：{err}。你可以手动输入模型名。")
        return gr.update(choices=[], value=None), f"拉取失败：{err}"

    if not models:
        gr.Warning("未探测到模型列表。请手动输入模型名后保存。")
        return gr.update(choices=[], value=None), "未探测到模型，请手动输入。"

    return gr.update(
        choices=models, value=models[0]
    ), f"拉取成功，共 {len(models)} 个模型"


def on_save_provider(
    provider_for_edit: str,
    provider_name: str,
    api_type_label: str,
    base_url: str,
    api_key: str,
    selected_model: str,
):
    name = (provider_name or "").strip()
    model_name = (selected_model or "").strip()
    if not name:
        return (
            gr.update(),
            gr.update(),
            gr.update(),
            "保存失败：提供商名称不能为空",
        )
    if not model_name:
        return (
            gr.update(),
            gr.update(),
            gr.update(),
            "保存失败：请先选择或输入模型名称",
        )

    api_type = LABEL_TO_API_TYPE.get(api_type_label, "openai_compatible")
    existing = CONFIG_MANAGER.get_provider(name)
    models: List[str] = []
    if existing and isinstance(existing.get("models"), list):
        models.extend(
            [str(m).strip() for m in existing.get("models", []) if str(m).strip()]
        )
    if model_name not in models:
        models.append(model_name)

    if provider_for_edit != "新增提供商" and provider_for_edit != name:
        CONFIG_MANAGER.delete_provider(provider_for_edit)

    CONFIG_MANAGER.upsert_provider(
        provider_name=name,
        api_type=api_type,
        base_url=(base_url or "").strip(),
        api_key=(api_key or "").strip(),
        models=models,
        default_model=model_name,
    )

    task_choices, task_default = get_task_choices()
    provider_choices = get_provider_selector_choices()
    return (
        gr.update(choices=task_choices, value=task_default),
        gr.update(choices=provider_choices, value=name),
        gr.update(choices=models, value=model_name),
        f"保存成功：{name}:{model_name}",
    )


with gr.Blocks(title="Vocabulary Pipeline") as demo:
    with gr.Tabs():
        with gr.Tab("任务看板"):
            with gr.Row():
                with gr.Column(scale=1):
                    choices, default_choice = get_task_choices()
                    provider_model_dropdown = gr.Dropdown(
                        label="模型配置（提供商:模型）",
                        choices=choices,
                        value=default_choice if default_choice else None,
                        allow_custom_value=False,
                    )
                    pdf_file = gr.File(
                        label="上传电子书",
                        file_types=[".pdf", ".epub", ".txt"],
                        type="filepath",
                    )
                    min_level = gr.Slider(
                        label="最低词汇等级 (如雅思分数)",
                        minimum=0,
                        maximum=9,
                        step=1,
                        value=0,
                    )
                    extract_mode_label = gr.Radio(
                        label="提取模式",
                        choices=["全部识别", "仅识别单词", "仅识别固定搭配"],
                        value="全部识别",
                    )
                    with gr.Row():
                        include_meaning = gr.Checkbox(
                            label="📜 导出语境释义", value=True
                        )
                        include_sentence = gr.Checkbox(
                            label="📖 导出原文句子", value=True
                        )
                    color_theme = gr.Radio(
                        ["黑白公文 (打印专用)", "护眼晨雾 (低饱和彩)"],
                        label="🎨 排版主题",
                        value="黑白公文 (打印专用)",
                    )
                    page_orientation = gr.Radio(
                        ["自动 (有原句则横向)", "竖向", "横向"],
                        label="🧭 页面方向",
                        value="自动 (有原句则横向)",
                    )
                    output_formats = gr.CheckboxGroup(
                        ["Excel (.xlsx)", "PDF (.pdf)", "TXT 单词本 (适配墨墨/欧路)"],
                        label="💾 导出格式",
                        value=["Excel (.xlsx)"],
                    )
                    export_preview = gr.Markdown(
                        summarize_output_formats(["Excel (.xlsx)"])
                    )
                    filter_strategy_choice = gr.Radio(
                        label="过滤策略",
                        choices=["专业阅读模式 (推荐)", "雅思备考模式 (白名单)"],
                        value="专业阅读模式 (推荐)",
                    )
                    gr.Markdown(
                        "**自定义词库（可选）说明**：上传后将覆盖内置策略。"
                        "若为 `word,level` 两列默认按白名单处理；"
                        "若文件名包含 black 将自动按黑名单处理。"
                    )
                    csv_file = gr.File(
                        label="自定义词库 (可选，将覆盖内置词库)", type="filepath"
                    )
                    start_button = gr.Button("启动任务", variant="primary")
                    stop_button = gr.Button("停止任务", variant="stop")

                    output_formats.change(
                        fn=summarize_output_formats,
                        inputs=[output_formats],
                        outputs=[export_preview],
                    )

                with gr.Column(scale=1):
                    log_box = gr.Textbox(label="运行日志", lines=28, interactive=False)
                    output_files = gr.File(
                        label="下载文件", file_count="multiple", interactive=False
                    )

            run_event = start_button.click(
                fn=process_pipeline,
                inputs=[
                    provider_model_dropdown,
                    min_level,
                    extract_mode_label,
                    include_meaning,
                    include_sentence,
                    color_theme,
                    page_orientation,
                    output_formats,
                    filter_strategy_choice,
                    pdf_file,
                    csv_file,
                ],
                outputs=[log_box, output_files],
            )
            stop_button.click(fn=None, cancels=[run_event])

        with gr.Tab("模型配置中心"):
            provider_for_edit = gr.Dropdown(
                label="选择提供商",
                choices=get_provider_selector_choices(),
                value="新增提供商",
            )
            provider_name = gr.Textbox(
                label="提供商名称", placeholder="例如: my-provider"
            )
            api_type = gr.Dropdown(
                label="接口类型",
                choices=list(LABEL_TO_API_TYPE.keys()),
                value="兼容OpenAI/第三方",
            )
            base_url = gr.Textbox(
                label="Base URL", placeholder="例如: https://api.openai.com/v1"
            )
            api_key = gr.Textbox(label="API Key", type="password", value="")
            fetch_models_button = gr.Button("🔄 测试并拉取模型")
            models_dropdown = gr.Dropdown(
                label="模型列表（可手动输入）",
                choices=[],
                value=None,
                allow_custom_value=True,
            )
            save_button = gr.Button("💾 保存此配置", variant="primary")
            config_status = gr.Textbox(label="配置状态", interactive=False)

            provider_for_edit.change(
                fn=on_provider_for_edit_change,
                inputs=[provider_for_edit],
                outputs=[
                    provider_name,
                    api_type,
                    base_url,
                    api_key,
                    models_dropdown,
                    config_status,
                ],
            )

            fetch_models_button.click(
                fn=on_fetch_models,
                inputs=[api_type, base_url, api_key],
                outputs=[models_dropdown, config_status],
            )

            save_button.click(
                fn=on_save_provider,
                inputs=[
                    provider_for_edit,
                    provider_name,
                    api_type,
                    base_url,
                    api_key,
                    models_dropdown,
                ],
                outputs=[
                    provider_model_dropdown,
                    provider_for_edit,
                    models_dropdown,
                    config_status,
                ],
            )


if __name__ == "__main__":
    demo.queue().launch(server_name="0.0.0.0", server_port=7860)
